from flask import Flask, render_template, request, redirect
import requests
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

ZAPIER_WEBHOOK_URL = https://hooks.zapier.com/hooks/catch/23360947/uy09xop/"
EXCEL_FILE = "reponses.xlsx"

app = Flask(__name__)

# Création du fichier Excel si inexistant
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Date", "Nom", "Prénom", "Adresse", "Code postal", "Quartier", "Téléphone",
        "Assister événements", "Participer événements quartier",
        "Intervenir voisins", "Rejoindre GA"
    ])
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def questionnaire():
    if request.method == "POST":
        data = {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nom": request.form.get("nom"),
            "prenom": request.form.get("prenom"),
            "adresse": request.form.get("adresse"),
            "code_postal": request.form.get("code_postal"),
            "quartier": request.form.get("quartier"),
            "telephone": request.form.get("telephone"),
            "assister_evenements": request.form.get("assister_evenements"),
            "participer_evenements": request.form.get("participer_evenements"),
            "intervenir_voisins": request.form.get("intervenir_voisins"),
            "rejoindre_ga": request.form.get("rejoindre_ga"),
            "commentaire": request.form.get("commentaire")
        }

        # Enregistrement Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(list(data.values()))
        wb.save(EXCEL_FILE)

        # Envoi vers Zapier
        try:
            requests.post(ZAPIER_WEBHOOK_URL, json=data)
        except Exception as e:
            print("Erreur Zapier :", e)

        return redirect("/merci")

    return render_template("formulaire.html")

@app.route("/merci")
def merci():
    return "<h2 style='text-align:center;'>✅ Merci, vos réponses ont bien été enregistrées.</h2>"

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=10000)

