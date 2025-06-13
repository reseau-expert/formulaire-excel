import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Création du fichier Excel s'il n'existe pas
excel_file = "reponses.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Nom", "Prénom", "Adresse", "Ville", "Code postal", "Rue ou résidence", "Ancienneté",
        "Ce que vous aimez", "À améliorer", "Activité souhaitée", "Type d’événements",
        "Souhaitez aider", "Besoin d’aide", "Personnes isolées connues",
        "Préférences de communication", "Groupes locaux", "Personnes relais",
        "Avis projets ville", "Jeunes - Manques", "Seniors - Besoins",
        "Idée de journée quartier", "Téléphone", "Email", "Commentaire résident", "Commentaire prospecteur"
    ])
    wb.save(excel_file)

# Questions du formulaire
questions = [
    "Votre nom", "Votre prénom", "Votre adresse complète", "Votre ville", "Votre code postal",
    "Dans quelle rue ou résidence habitez-vous ?", "Depuis combien de temps vivez-vous ici ?",
    "Qu’est-ce qui vous plaît le plus dans votre quartier ?", "Si vous deviez améliorer UNE chose ici, ce serait quoi ?",
    "Quelle activité aimeriez-vous faire dans le quartier ?",
    "Préférez-vous des événements : ludiques, utiles ou éducatifs ?",
    "Auriez-vous besoin d’un coup de main pour quelque chose ?",
    "Seriez-vous prêt à aider un voisin occasionnellement ?",
    "Connaissez-vous des personnes isolées à aider ?",
    "Comment préférez-vous recevoir les infos ?", "Faites-vous partie d’un groupe local ?",
    "Qui dans votre entourage serait intéressé ?", "Souhaitez-vous donner votre avis sur les projets ?",
    "Jeunes : Que manque-t-il pour vous ?", "Seniors : Qu’est-ce qui faciliterait votre quotidien ?",
    "Si on organisait une journée ‘quartier idéal’, que proposeriez-vous ?",
    "Votre numéro de téléphone", "Votre adresse email",
    "Commentaire du résident", "Commentaire du prospecteur"
]

# Interface graphique
root = tk.Tk()
root.title("Questionnaire Citoyen")

canvas = tk.Canvas(root, borderwidth=0)
frame = tk.Frame(canvas)
vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=vsb.set)

vsb.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)
canvas.create_window((0, 0), window=frame, anchor="n")

entries = []
for q in questions:
    label = tk.Label(frame, text=q, font=("Helvetica", 10), anchor="center", justify="center")
    label.pack(pady=(10, 0))
    entry = tk.Entry(frame, width=80)
    entry.pack(padx=10, pady=(0, 10))
    entries.append(entry)

def onFrameConfigure(canvas):
    canvas.configure(scrollregion=canvas.bbox("all"))

frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))

def submit():
    values = [entry.get() for entry in entries]
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append(values)
        wb.save(excel_file)
        messagebox.showinfo("Succès", "Vos réponses ont bien été enregistrées.")
        for entry in entries:
            entry.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur est survenue : {e}")

submit_btn = tk.Button(frame, text="Envoyer", command=submit, bg="green", fg="white", font=("Helvetica", 12, "bold"))
submit_btn.pack(pady=20)

root.mainloop()